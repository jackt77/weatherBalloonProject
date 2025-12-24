import csv
import requests
import datetime
from io import StringIO

def fetch_predictions(
	launch_position: dict[str, float],
	launch_altitude: float,
	ascent_rate: float,
	burst_altitude: float,
	descent_rate: float
) -> dict[int, list[dict]]:
	"""
	Fetch a flight prediction from the SondeHub Tawhiri API and return the CSV
	data as a list of dictionaries.
	Parameters:
	    launch_position (dict[str, float]): Launch position as {"lat": float, "lon": float}.
	    launch_altitude (float): Launch altitude (m)
	    ascent_rate (float): Ascent rate (ms⁻¹)
	    burst_altitude (float): Burst altitude (m)
	    descent_rate (float): Descent rate (ms⁻¹)
	Returns:
	     Predictions as dict[int, list[dict]]
	"""
	predictions = {}
	for hour in [0, 6, 12, 18]:
		launch_time = datetime.datetime.now().replace(hour=hour, minute=0, second=0, microsecond=0)
		params = {
			"profile": "standard_profile",
			"pred_type": "single",
			"launch_datetime": launch_time.strftime("%Y-%m-%dT%H:%M:%SZ"),
			"launch_latitude": launch_position["lat"],
			"launch_longitude": launch_position["lon"],
			"launch_altitude": launch_altitude,
			"ascent_rate": ascent_rate,
			"burst_altitude": burst_altitude,
			"descent_rate": descent_rate,
			"format": "csv",
		}
		response = requests.get("https://api.v2.sondehub.org/tawhiri", params=params)
		response.raise_for_status()
		predictions[hour] = list(csv.DictReader(StringIO(response.text)))
	return predictions

def get_date_from_datetime(date_time: str) -> str:
	date = date_time[:10]
	return date

def get_time_from_datetime(date_time: str) -> str:
	time = date_time[10:].strip().replace("T", """""").replace("Z", """""")
	return time

import openpyxl
import pathlib
from openpyxl.utils import get_column_letter

def save_predictions_to_excel(predictions: dict[int, list[dict]]) -> None:
	entry_number = 0
	date: str = get_date_from_datetime(predictions[0][0]["datetime"])
	base_path = pathlib.Path(__file__).parent
	file_path = base_path / "data.xlsx"
	workbook = openpyxl.open(file_path)
	worksheet = workbook.create_sheet(title=f"{date}")
	headers = ["time", "latitude", "longitude", "altitude"]
	for hour, prediction in predictions.items():
		offset: int = (7 * entry_number)
		for col, header in enumerate(headers, start=1):
			worksheet.cell(row=1, column=offset + col).value = header
		for row, entry in enumerate(prediction, start=2):
			time: str = get_time_from_datetime(entry["datetime"])
			worksheet.cell(row=row, column=offset + 1).value = time
			worksheet.cell(row=row, column=offset + 2).value = float(entry["latitude"])
			worksheet.cell(row=row, column=offset + 3).value = float(entry["longitude"])
			worksheet.cell(row=row, column=offset + 4).value = float(entry["altitude"])
		for col in range(1, 5):
			max_length = 0
			col_letter = get_column_letter(offset + col)
			for cell in worksheet[col_letter]:
				value = str(cell.value)
				max_length = max(max_length, len(value))
			worksheet.column_dimensions[col_letter].width = max_length + 2
		entry_number += 1
		workbook.save(file_path)

def main():
	# Default values
		# launch_position = {"lat":54.66864, "lon":356.6498}    Cockermouth school astro coordinates
		# launch_altitude = 81.0                                Altitude on the astro
		# ascent_rate = 6.0
		# burst_altitude = 35000                                Based on kaymont balloon's specs
		# descent_rate = 6.0
	data = fetch_predictions(
		{"lat":54.66864, "lon":356.6498},
		81.0,
		6.0,
		30000,
		6.0
	)
	save_predictions_to_excel(data)

if __name__ == "__main__":
	main()
